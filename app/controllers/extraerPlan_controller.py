import datetime
import os
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog
from app.models.database import create_connection
from app.models.gestion_model.docente_model import DocenteModel
from app.models.gestion_model.contrato_model import ContratoModel
from app.models.gestion_model.accion_model import AccionModel

class ExtraerPlanController:
    def __init__(self):
        self.view = None
        self.docente_model = DocenteModel()
        self.contrato_model = ContratoModel()
        self.accion_model = AccionModel()

    def show_view(self):
        if not self.view:
            from app.views.extraerPlan_view import ExtraerPlanView
            self.view = ExtraerPlanView(self)
            self.view.deiconify()

    def process_excel_files(self):
        folder_path = filedialog.askdirectory(title="Selecciona la carpeta con los archivos Excel")

        if not folder_path:
            messagebox.showerror("Error", "No se seleccionó ninguna carpeta.")
            return

        if not os.path.exists(folder_path):
            messagebox.showerror("Error", f"La carpeta especificada no existe: {folder_path}")
            return

        conn = create_connection()
        total_files = 0
        extracted_data = []

        try:
            for filename in os.listdir(folder_path):
                if filename.endswith('.xlsx'):
                    total_files += 1
                    filepath = os.path.join(folder_path, filename)
                    data = self.process_excel_file(filepath, conn)
                    if data:
                        extracted_data.append(data)

            messagebox.showinfo("Procesado", f"Se han procesado {total_files} archivos exitosamente.")
            self.view.update_info(total_files, extracted_data)
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar los archivos:\n{str(e)}")
        finally:
            conn.close()

    def process_excel_file(self, filepath, conn):
        try:
            # Abrir el archivo Excel con openpyxl
            workbook = openpyxl.load_workbook(filepath)
            
            # Acceder a las hojas deseadas
            if 'Hoja1' not in workbook.sheetnames or 'Hoja2' not in workbook.sheetnames:
                messagebox.showerror("Error", f"El archivo {os.path.basename(filepath)} no contiene las hojas necesarias.")
                return None
            
            sheet1 = workbook['Hoja1']
            sheet2 = workbook['Hoja2']
            
            """# Ejemplo de cómo manejar una celda combinada (A6:H6)
            cell_value = None
            for row in sheet1.iter_rows(min_row=6, max_row=6, min_col=1, max_col=8):
                for cell in row:
                    if cell.value:
                        cell_value = cell.value
                        break
                if cell_value:
                    break"""
            
            # Construir el diccionario de datos del docente
            docente_data = {
                "nombre_docente": sheet1['A8'].value,
                "correo_docente": sheet1['U6'].value,
                "dni_docente": sheet1['P8'].value,
                "celular_docente": sheet1['AB8'].value,
                "gradoBachiller_docente": sheet2['D2'].value,
                "tituloProfesional_docente": sheet2['D3'].value,
                "gradoMaestro_docente": sheet2['D4'].value,
                "gradoDoctor_docente": sheet2['D5'].value,
                "tituloSegEspMedico_docente": sheet2['D6'].value,
                "tituloSegEspOdontologo_docente": sheet2['D7'].value
            }
            
            # Insertar datos del docente en la base de datos
            docente_id = self.docente_model.insert_docente_desde_excel(docente_data)
            

            # Construir el diccionario de datos del contrato
            contrato_data = {
                "renacyt_contrato": sheet2['C13'].value,
                "id_condicion": self.get_id_from_table(conn, "Condicion", "nombre_condicion", sheet2['C11'].value),
                "id_regimen": self.get_id_from_table(conn, "Regimen", "nombre_regimen", sheet2['C9'].value),
                "id_categoria": self.get_id_from_table(conn, "Categoria", "nombre_categoria", sheet2['G9'].value),
                "id_docente": docente_id
            }
            
            # Insertar datos del contrato en la base de datos
            self.contrato_model.insert_contrato_desde_excel(contrato_data)
            


            # Iterar sobre las filas y columnas del rango especificado
            days_of_week = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO']
            start_row = 12
            end_row = 27
            start_col = 2  # Columna B
            end_col = 31   # Columna AE
            columns_per_day = 5
            
            actions = []
            
            for row in range(start_row, end_row + 1):
                hora_inicio = sheet1.cell(row=row, column=1).value  # Horas en la columna A
                hora_fin = sheet1.cell(row=row, column=1).value#(datetime.datetime.combine(datetime.date.today(), hora_inicio) + datetime.timedelta(hours=1)).time()  # Asumiendo que cada bloque es de 1 hora
                
                for day_index, day in enumerate(days_of_week):
                    day_col_start = start_col + day_index * columns_per_day
                    ambiente = sheet1.cell(row=row, column=day_col_start + 1).value
                    sigla_tipoActividad = sheet1.cell(row=row, column=day_col_start + 2).value
                    num_alumnos = sheet1.cell(row=row, column=day_col_start + 4).value
                    
                    if ambiente and sigla_tipoActividad and num_alumnos:  # Verificar si hay datos válidos en las celdas
                        action_data = {
                            "dia_accion": day,
                            "horaInicio_accion": hora_inicio,
                            "horaFin_accion": hora_fin,
                            "ambiente_accion": ambiente,
                            "numAlumnos_accion": num_alumnos,
                            "id_tipoActividad": self.get_id_from_table_Tipo_Actividad(conn, "Tipo_Actividad", "sigla_tipoActividad", sigla_tipoActividad),
                            "id_semana": self.get_id_from_table(conn, "Semana", "nombre_semana", 'SEMANA 1'),  # Asumiendo 'SEMANA 1' para simplificar
                            "id_docente": docente_id
                        }
                            #"id_regimen": self.get_id_from_table(conn, "Regimen", "nombre_regimen", sheet2['C9'].value),
                        
                        self.accion_model.insert_accion_desde_excel(action_data)
                        actions.append(action_data)
            return {
                "archivo": os.path.basename(filepath),
                "docente": docente_data,
                "contrato": contrato_data
            }
        except KeyError as e:
            messagebox.showerror("Error", f"Error al procesar el archivo {os.path.basename(filepath)}:\nColumna no encontrada: {str(e)}")
            return None

    def get_id_from_table(self, conn, table_name, column_name, value):
        cursor = conn.cursor()
        cursor.execute(f"SELECT id_{table_name.lower()} FROM {table_name} WHERE {column_name} = ?", (value,))
        result = cursor.fetchone()
        return result[0] if result else None
    
    def get_id_from_table_Tipo_Actividad(self, conn, table_name, column_name, value):
        cursor = conn.cursor()
        cursor.execute(f"SELECT id_tipoActividad FROM {table_name} WHERE {column_name} = ?", (value,))
        result = cursor.fetchone()
        return result[0] if result else None

